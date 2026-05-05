"""
Read acres from D15 in each generated invoice sheet,
match to project rows in the AEA form, and fill the Acres column.

Matching strategy (in priority order):
  1. Multi-sheet invoices: match sheet name to project title (normalized,
     truncated at 31 chars, special chars replaced with underscore/space)
  2. Single-sheet 'Template' invoices: match by PI email from filename
     to 'Faculty PI Email address' column in the form
  3. Single-sheet 'Template' invoices: match by PI last name from filename
     when email gives exactly one form row
"""

import re
from pathlib import Path
import openpyxl

INVOICES_DIR = Path(__file__).parent / "Generated_Invoices"
FORM_PATH = Path(__file__).parent / "AEA_2025_Research_Project_Form.xlsx"


def norm(s):
    return re.sub(r"\s+", " ", str(s or "").strip()).lower()


def norm_title(s):
    """Normalize a project title / sheet name for comparison.
    Collapses whitespace, lowercases, and replaces special chars with space."""
    s = str(s or "").strip()
    s = re.sub(r"[/:&+]", " ", s)        # colon/slash/ampersand → space
    s = re.sub(r"[_\-]", " ", s)          # underscore/dash → space
    s = re.sub(r"\s+", " ", s).lower().strip()
    return s


def sheet_base(name):
    """Strip trailing duplicate suffix like '1', '_1', '_2' from sheet names."""
    return re.sub(r"[_ ]\d+$", "", name).strip()


def match_title(sheet_name, project_title):
    """Return True if sheet_name plausibly matches the (possibly truncated) project_title."""
    sn = norm_title(sheet_base(sheet_name))
    pt = norm_title(project_title)
    # Truncate project title at 31 chars (Excel sheet name limit)
    pt_trunc = norm_title(project_title[:31].strip())
    return (
        sn == pt
        or sn == pt_trunc
        or pt.startswith(sn)
        or sn.startswith(pt_trunc)
    )


def parse_invoice_filename(stem):
    """Return (pi_full_name, email, pi_last_name_lower) from invoice filename stem."""
    parts = stem.split("_")
    pi_full = parts[0].strip()
    email = parts[1].strip() if len(parts) > 1 else ""
    pi_last = norm(pi_full.split()[-1]) if pi_full.split() else norm(pi_full)
    return pi_full, email, pi_last


def main():
    # --- Step 1: harvest invoice data ---
    # by_last[pi_last] = list of (sheet_name, acres, email, inv_path)
    # by_email[email]  = list of (sheet_name, acres, inv_path)
    by_last = {}
    by_email = {}

    for inv_path in sorted(INVOICES_DIR.glob("*.xlsx")):
        pi_full, email, pi_last = parse_invoice_filename(inv_path.stem)
        wb = openpyxl.load_workbook(inv_path, data_only=True)
        sheets = [(sn, wb[sn]["D15"].value) for sn in wb.sheetnames]
        wb.close()

        for sn, v in sheets:
            by_last.setdefault(pi_last, []).append((sn, v, email, inv_path))
        if email:
            for sn, v in sheets:
                by_email.setdefault(email.lower(), []).append((sn, v, inv_path))

    # --- Step 2: load form ---
    form_wb = openpyxl.load_workbook(FORM_PATH)
    form_ws = form_wb.active

    headers = [form_ws.cell(1, c).value for c in range(1, form_ws.max_column + 1)]
    col = {h: i + 1 for i, h in enumerate(headers)}
    proj_title_col = col["Project Title"]
    pi_last_col = col["Faculty PI Last Name"]
    pi_email_col = col["Faculty PI Email address"]
    acres_col = col["Acres"]

    # Index form rows
    pi_rows = {}     # pi_last_lower -> [row]
    email_rows = {}  # email_lower   -> [row]
    for row in range(2, form_ws.max_row + 1):
        pi_last = norm(form_ws.cell(row, pi_last_col).value)
        email = norm(form_ws.cell(row, pi_email_col).value)
        if pi_last:
            pi_rows.setdefault(pi_last, []).append(row)
        if email:
            email_rows.setdefault(email, []).append(row)

    updated = 0
    no_match = []
    already_set = set()  # (row) to avoid double-filling

    def write_acres(row, acres_val, reason):
        nonlocal updated
        if row in already_set:
            return
        already_set.add(row)
        form_ws.cell(row, acres_col).value = acres_val
        pt = form_ws.cell(row, proj_title_col).value
        print(f"  OK  row={row:3d} | {reason:55s} | {str(pt)[:40]:40s} | D15={acres_val}")
        updated += 1

    # --- Step 3a: match by PI last name (multi-sheet and single-sheet with unique PI) ---
    for pi_last, entries in by_last.items():
        rows = pi_rows.get(pi_last, [])

        # Separate real-named sheets from 'Template' ones
        real_entries = [(sn, v, em, p) for sn, v, em, p in entries if sn.lower() != "template"]
        tmpl_entries = [(sn, v, em, p) for sn, v, em, p in entries if sn.lower() == "template"]

        # Multi-sheet: match each form row by project title
        if real_entries and rows:
            for row in rows:
                pt = str(form_ws.cell(row, proj_title_col).value or "")
                for sn, acres_val, em, _ in real_entries:
                    if match_title(sn, pt):
                        write_acres(row, acres_val, f"last={pi_last!r} sheet={sn!r}")
                        break

        # Single-sheet Template with exactly one matching form row
        if tmpl_entries and len(rows) == 1:
            sn, acres_val, em, _ = tmpl_entries[0]
            write_acres(rows[0], acres_val, f"last={pi_last!r} single-template")

    # --- Step 3b: match remaining by PI email from invoice filename ---
    for email, entries in by_email.items():
        rows = email_rows.get(email, [])
        if not rows:
            continue

        real_entries = [(sn, v, p) for sn, v, p in entries if sn.lower() != "template"]
        tmpl_entries = [(sn, v, p) for sn, v, p in entries if sn.lower() == "template"]

        if real_entries and rows:
            for row in rows:
                if row in already_set:
                    continue
                pt = str(form_ws.cell(row, proj_title_col).value or "")
                for sn, acres_val, _ in real_entries:
                    if match_title(sn, pt):
                        write_acres(row, acres_val, f"email={email} sheet={sn!r}")
                        break

        if tmpl_entries and len(rows) == 1:
            sn, acres_val, _ = tmpl_entries[0]
            write_acres(rows[0], acres_val, f"email={email} single-template")

    # --- Report remaining missing ---
    form_wb.save(FORM_PATH)
    print(f"\nSaved. Updated {updated} rows.")

    missing = []
    for row in range(2, form_ws.max_row + 1):
        pi_last = norm(form_ws.cell(row, pi_last_col).value)
        if pi_last and form_ws.cell(row, acres_col).value is None:
            pt = form_ws.cell(row, proj_title_col).value
            missing.append((row, pi_last, pt))
    if missing:
        print(f"\nStill missing acres ({len(missing)}):")
        for r, p, t in missing:
            print(f"  row {r:3d}: {p!r:25s} | {t!r}")


if __name__ == "__main__":
    main()
