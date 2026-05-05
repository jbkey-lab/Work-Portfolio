import argparse
import csv
import re
import shutil
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

DEFAULT_OUTPUT = Path.home() / "Documents" / "AEAInvoice" / "Generated_Invoices"
E12_MAX_CHARS = 31


def _safe_str(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate individualized project leader invoices from template."
    )
    parser.add_argument("--input", required=True, help="Path to the research project form (.csv or .xlsx).")
    parser.add_argument("--bills", required=True, help="Path to bills.xlsx (must be .xlsx).")
    parser.add_argument("--template", required=True, help="Path to PI_Year_AEA.xlsx template (must be .xlsx).")
    parser.add_argument(
        "--output",
        default=str(DEFAULT_OUTPUT),
        help=f"Output directory (default: {DEFAULT_OUTPUT}).",
    )
    parser.add_argument("--year", help="Optional ReportingYear filter.")
    return parser.parse_args(argv)


def _read_csv_rows(input_path: Path) -> list[dict]:
    with input_path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        return [{_safe_str(k).lower(): _safe_str(v) for k, v in row.items()} for row in reader]


def _read_xlsx_rows(input_path: Path) -> list[dict]:
    wb = load_workbook(input_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    headers = [_safe_str(h).lower() for h in rows[0]]
    records = []
    for row_vals in rows[1:]:
        record = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            val = row_vals[idx] if idx < len(row_vals) else None
            record[header] = _safe_str(val)
        records.append(record)
    return records


def _load_rows(input_path: Path) -> list[dict]:
    if input_path.suffix.lower() == ".csv":
        return _read_csv_rows(input_path)
    if input_path.suffix.lower() == ".xlsx":
        return _read_xlsx_rows(input_path)
    raise ValueError("Only .xlsx and .csv inputs are supported")


def _filter_year(rows: list[dict], year: str | None) -> list[dict]:
    if not year:
        return list(rows)
    return [row for row in rows if _safe_str(row.get("reportingyear")) == str(year)]


def _load_bills_lookup(bills_path: Path) -> dict:
    lookup: dict = {}
    for row in _read_xlsx_rows(bills_path):
        full_name = _safe_str(row.get("projectleader"))
        if not full_name:
            continue
        parts = full_name.split()
        last_name = parts[-1].lower() if parts else ""
        if last_name and last_name not in lookup:
            lookup[last_name] = {
                "full_name": full_name,
                "email": _safe_str(row.get("email")),
                "dept": _safe_str(row.get("department")),
                "id": _safe_str(row.get("id")),
            }
    return lookup


def _project_title_report(title: str) -> str:
    if len(title) <= E12_MAX_CHARS:
        return title
    cut = title[:E12_MAX_CHARS]
    last_space = cut.rfind(" ")
    return cut[:last_space].rstrip() if last_space > 0 else cut


def _read_unit_prices(template_path: Path) -> dict[int, float]:
    wb = load_workbook(template_path, data_only=True)
    ws = wb.active
    prices: dict[int, float] = {}
    for row_num in range(15, 30):
        val = ws.cell(row=row_num, column=5).value  # column E
        if val is not None:
            try:
                prices[row_num] = float(val)
            except (ValueError, TypeError):
                pass
    wb.close()
    return prices


def _project_total(acres_str: str, unit_prices: dict[int, float]) -> float | None:
    s = acres_str.strip()
    if not s:
        return None
    try:
        return round(float(s) * unit_prices.get(15, 336.0), 2)
    except ValueError:
        return None


def _write_summary_csv(output_dir: Path, summary: list[dict]) -> None:
    csv_path = output_dir / "summary.csv"
    fieldnames = ["PI Name", "PI Email", "Project Title", "Acres", "Total ($)", "Project Contact Name", "Project Contact Email"]
    with csv_path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        grand_total = 0.0
        grand_acres = 0.0
        for entry in summary:
            pi_total = 0.0
            pi_acres = 0.0
            for proj in entry["projects"]:
                acres_str = proj["acres_str"]
                total_val = proj["total_val"]
                try:
                    acres_disp = f"{float(acres_str):.4g}" if acres_str else ""
                except (ValueError, TypeError):
                    acres_disp = acres_str or ""
                writer.writerow({
                    "PI Name": entry["pi_name"],
                    "PI Email": entry["email"],
                    "Project Title": proj["title"],
                    "Acres": acres_disp,
                    "Total ($)": f"{total_val:.2f}" if total_val is not None else "TBD",
                    "Project Contact Name": proj.get("contact_name", ""),
                    "Project Contact Email": proj.get("contact_email", ""),
                })
                if total_val is not None:
                    pi_total += total_val
                try:
                    pi_acres += float(acres_str)
                except (ValueError, TypeError):
                    pass
            writer.writerow({
                "PI Name": f"{entry['pi_name']} SUBTOTAL",
                "PI Email": "",
                "Project Title": "",
                "Acres": f"{pi_acres:.2f}",
                "Total ($)": f"{pi_total:.2f}",
                "Project Contact Name": "",
                "Project Contact Email": "",
            })
            grand_total += pi_total
            grand_acres += pi_acres
        writer.writerow({
            "PI Name": "GRAND TOTAL",
            "PI Email": "",
            "Project Title": "",
            "Acres": f"{grand_acres:.2f}",
            "Total ($)": f"{grand_total:.2f}",
            "Project Contact Name": "",
            "Project Contact Email": "",
        })
    print(f"Summary written: {csv_path}")


def _safe_clear_output_dir(output_dir: Path, input_paths: list[Path]) -> None:
    if output_dir.exists():
        resolved_out = output_dir.resolve()
        for p in input_paths:
            try:
                p.resolve().relative_to(resolved_out)
                print(
                    f"Error: output directory '{output_dir}' contains input file '{p.name}'. "
                    "Specify a different --output directory to avoid deleting source files."
                )
                sys.exit(2)
            except ValueError:
                pass
        locked = [f for f in output_dir.iterdir() if f.name.startswith("~$")]
        if locked:
            print(
                f"Error: the following files in '{output_dir}' are open in Excel. "
                "Close them and re-run:\n" + "\n".join(f"  {f.name}" for f in locked)
            )
            sys.exit(2)
        shutil.rmtree(output_dir)
    output_dir.mkdir(parents=True)


def _safe_sheet_name(title: str, seen: dict[str, int]) -> str:
    safe = re.sub(r"[\\/*?:\[\]]", "_", title)[:31]
    if safe in seen:
        seen[safe] += 1
        suffix = f"_{seen[safe]}"
        safe = safe[: 31 - len(suffix)] + suffix
    else:
        seen[safe] = 1
    return safe


def _write_invoice(
    template_path: Path,
    dest_path: Path,
    projects: list[tuple[str, str, str, str]],  # (sheet_name, pi_full_name, acres_str, title)
    dept: str,
) -> None:
    wb = load_workbook(template_path)
    base_ws = wb.active

    # Copy the clean template sheet before filling anything
    while len(wb.worksheets) < len(projects):
        wb.copy_worksheet(base_ws)

    seen: dict[str, int] = {}
    for i, (sheet_name, pi_full_name, acres_str, title) in enumerate(projects):
        ws = wb.worksheets[i]
        ws["B3"] = pi_full_name
        ws["B7"] = pi_full_name
        ws["B8"] = dept
        ws["E10"] = pi_full_name
        ws["E12"] = _project_title_report(title)
        try:
            ws["D15"] = float(acres_str) if acres_str.strip() else None
        except ValueError:
            ws["D15"] = None
        ws.title = _safe_sheet_name(sheet_name, seen)

    wb.calculation.fullCalcOnLoad = True
    wb.save(dest_path)


def generate_invoices(
    input_path: Path, bills_path: Path, template_path: Path, output_dir: Path, year: str | None
) -> None:
    _safe_clear_output_dir(output_dir, [input_path, bills_path, template_path])

    input_rows = _filter_year(_load_rows(input_path), year)
    bills_lookup = _load_bills_lookup(bills_path)
    unit_prices = _read_unit_prices(template_path)

    groups: dict[str, list[dict]] = defaultdict(list)
    for row in input_rows:
        last_name = _safe_str(row.get("faculty pi last name")).strip()
        if last_name:
            groups[last_name.lower()].append(row)

    summary: list[dict] = []

    for last_name_key, rows in groups.items():
        info = bills_lookup.get(last_name_key, {})
        pi_full_name = info.get("full_name") or rows[0].get("faculty pi last name", last_name_key)
        email = info.get("email", "")
        dept = info.get("dept", "Unknown Department")
        reporting_year = _safe_str(rows[0].get("reportingyear"))

        safe_name = re.sub(r'[<>:"/\\|?*]', "_", pi_full_name)
        filename = f"{safe_name}_{reporting_year}_AEA.xlsx"
        dest_path = output_dir / filename

        projects: list[tuple] = []
        pi_projects: list[dict] = []
        for row in rows:
            raw_title = _safe_str(row.get("short project title")) or _safe_str(row.get("project title"))
            acres = _safe_str(row.get("acres"))
            sheet_name = _project_title_report(raw_title) if raw_title else "Project"
            projects.append((sheet_name, pi_full_name, acres, sheet_name))
            pi_projects.append({
                "title": sheet_name,
                "acres_str": acres,
                "total_val": _project_total(acres, unit_prices),
                "contact_name": _safe_str(row.get("project contact name")),
                "contact_email": _safe_str(row.get("project contact email address")),
            })

        _write_invoice(template_path, dest_path, projects, dept)
        summary.append({"pi_name": pi_full_name, "email": email, "projects": pi_projects})

    _write_summary_csv(output_dir, summary)


def main(argv: list[str]) -> int:
    args = _parse_args(argv)
    input_path = Path(args.input)
    bills_path = Path(args.bills)
    template_path = Path(args.template)
    output_dir = Path(args.output)

    for p in [input_path, bills_path, template_path]:
        if not p.exists():
            print(f"File not found: {p}")
            return 2

    if bills_path.suffix.lower() != ".xlsx" or template_path.suffix.lower() != ".xlsx":
        print("Error: The --bills and --template files must be in .xlsx format.")
        return 2

    try:
        generate_invoices(input_path, bills_path, template_path, output_dir, args.year)
    except Exception as exc:
        import traceback
        print(f"Invoice generation failed: {exc}")
        traceback.print_exc()
        return 1

    print(f"Individualized files generated in: {output_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
