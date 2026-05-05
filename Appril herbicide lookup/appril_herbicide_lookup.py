from __future__ import annotations

import argparse
import json
import re
import sys
import time
from collections import defaultdict
from dataclasses import dataclass
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any, Iterable
from urllib.error import HTTPError, URLError
from urllib.parse import urlencode
from urllib.request import Request, urlopen

from openpyxl import Workbook, load_workbook


API_URL = "https://ordspub.epa.gov/ords/pesticides/apprilapi/"
DEFAULT_DUMP_URL = "https://www3.epa.gov/pesticides/appril/apprildatadump_public.xlsx"
DEFAULT_HEADER_CANDIDATES = ("herbicide", "product", "product name", "name")


@dataclass
class CandidateMatch:
    input_name: str
    primary_name: str
    search_term: str
    score: float
    record: dict[str, Any]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Read an Excel list of herbicides, search the EPA APPRIL API, and "
            "write a workbook with best matches, all possible matches, and "
            "ingredient-group duplicate hints."
        )
    )
    parser.add_argument("input_file", type=Path, help="Path to the source Excel workbook")
    parser.add_argument(
        "-o",
        "--output-file",
        type=Path,
        help="Path to the output workbook. Defaults to '<input>_appril_matches.xlsx'",
    )
    parser.add_argument("--sheet", help="Sheet name to read. Defaults to the first worksheet.")
    parser.add_argument(
        "--column",
        help="Header name for the herbicide column. Defaults to auto-detect or first column.",
    )
    parser.add_argument(
        "--include-inactive",
        action="store_true",
        help="Include inactive/cancelled products instead of filtering to active products only.",
    )
    parser.add_argument(
        "--min-score",
        type=float,
        default=0.35,
        help="Discard matches below this similarity score. Default: 0.35",
    )
    parser.add_argument(
        "--per-term-limit",
        type=int,
        default=100,
        help="Maximum APPRIL rows to fetch per search term. Default: 100",
    )
    parser.add_argument(
        "--sleep-seconds",
        type=float,
        default=0.15,
        help="Delay between API calls to avoid hammering the service. Default: 0.15",
    )
    parser.add_argument(
        "--source",
        choices=("auto", "dump", "api"),
        default="auto",
        help="Lookup source. 'auto' prefers the local dump when provided. Default: auto",
    )
    parser.add_argument(
        "--appril-dump",
        type=Path,
        help="Path to the APPRIL public data dump workbook. Recommended because the live API may be unavailable.",
    )
    return parser.parse_args()


def load_input_rows(
    input_file: Path,
    sheet_name: str | None,
    primary_column_name: str | None,
    search_column_name: str | None,
) -> list[dict[str, str]]:
    workbook = load_workbook(input_file, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
        rows = worksheet.iter_rows(values_only=True)
        header_row = next(rows, None)
        if header_row is None:
            return []

        header_values = [str(value).strip() if value is not None else "" for value in header_row]
        primary_index = pick_column_index(header_values, primary_column_name)
        search_index = pick_column_index(header_values, search_column_name, allow_missing=True)
        epa_index = pick_column_index(header_values, "EPA  Number", allow_missing=True)
        if epa_index is None:
            epa_index = pick_column_index(header_values, "EPA Number", allow_missing=True)

        records: list[dict[str, str]] = []
        for row in rows:
            if primary_index >= len(row):
                continue
            primary_value = row[primary_index]
            if primary_value is None:
                continue
            primary_text = str(primary_value).strip()
            if not primary_text:
                continue

            search_text = primary_text
            if search_index is not None and search_index < len(row):
                search_value = row[search_index]
                if search_value is not None and str(search_value).strip():
                    search_text = str(search_value).strip()

            epa_number = ""
            if epa_index is not None and epa_index < len(row):
                epa_value = row[epa_index]
                if epa_value is not None and str(epa_value).strip():
                    epa_number = str(epa_value).strip()

            records.append(
                {
                    "primary_name": primary_text,
                    "search_term": search_text,
                    "epa_number": epa_number,
                }
            )
        return records
    finally:
        workbook.close()


def pick_column_index(headers: list[str], column_name: str | None, allow_missing: bool = False) -> int | None:
    lowered = [header.strip().lower() for header in headers]
    if column_name:
        target = column_name.strip().lower()
        if target in lowered:
            return lowered.index(target)
        if allow_missing:
            return None
        raise ValueError(f"Column '{column_name}' was not found in the header row: {headers}")

    for candidate in DEFAULT_HEADER_CANDIDATES:
        if candidate in lowered:
            return lowered.index(candidate)

    for index, value in enumerate(headers):
        if value.strip():
            return index

    return None if allow_missing else 0


def build_search_terms(name: str) -> list[str]:
    base = normalize_spacing(name)
    return [base] if base else []


def remove_common_noise(text: str) -> str:
    noise_words = {
        "herbicide",
        "plus",
        "ii",
        "iii",
        "iv",
        "ultra",
        "max",
        "maxx",
        "extra",
    }
    kept = [word for word in text.split() if word.lower() not in noise_words]
    return normalize_spacing(" ".join(kept))


def normalize_spacing(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def search_appril(term: str, include_inactive: bool, per_term_limit: int, sleep_seconds: float) -> list[dict[str, Any]]:
    query: dict[str, Any] = {
        "$or": [
            {"product_name": {"$instr": term}},
            {"abns": {"$instr": term}},
        ]
    }
    if not include_inactive:
        query["status_group"] = "Active"

    items: list[dict[str, Any]] = []
    offset = 0
    page_size = min(per_term_limit, 500)

    while len(items) < per_term_limit:
        response = fetch_json(
            {
                "q": json.dumps(query, separators=(",", ":")),
                "limit": str(page_size),
                "offset": str(offset),
            }
        )
        page_items = response.get("items", [])
        if not isinstance(page_items, list):
            break

        items.extend(item for item in page_items if isinstance(item, dict))

        has_more = bool(response.get("hasMore"))
        if not has_more or not page_items:
            break

        offset += len(page_items)
        if sleep_seconds > 0:
            time.sleep(sleep_seconds)

    return items[:per_term_limit]


def load_dump_records(dump_path: Path, include_inactive: bool) -> list[dict[str, Any]]:
    workbook = load_workbook(dump_path, read_only=True, data_only=True)
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        rows = worksheet.iter_rows(values_only=True)
        raw_headers = next(rows, None)
        if raw_headers is None:
            return []

        headers = [str(value).strip().upper() if value is not None else "" for value in raw_headers]
        records: list[dict[str, Any]] = []
        for row in rows:
            raw_record = {
                headers[index]: row[index]
                for index in range(min(len(headers), len(row)))
                if headers[index]
            }
            status_group = stringify_value(raw_record.get("STATUS_GROUP"))
            if not include_inactive and status_group.lower() != "active":
                continue
            records.append(
                {
                    "reg_num": stringify_value(raw_record.get("REG_NUM")),
                    "product_name": stringify_value(raw_record.get("PRODUCT_NAME")),
                    "abns": stringify_value(raw_record.get("ABNS")),
                    "ais": stringify_value(raw_record.get("AIS")),
                    "status_group": status_group,
                    "reg_type": stringify_value(raw_record.get("REG_TYPE")),
                    "company_name": stringify_value(raw_record.get("COMPANY_NAME")),
                }
            )
        return records
    finally:
        workbook.close()


def search_dump(term: str, dump_records: list[dict[str, Any]], per_term_limit: int) -> list[dict[str, Any]]:
    normalized_term = normalize_text(term)
    if not normalized_term:
        return []

    exact_matches: list[tuple[float, dict[str, Any]]] = []
    for record in dump_records:
        product_name = normalize_text(get_field(record, "product_name"))
        abns_value = get_field(record, "abns")
        normalized_abns = normalize_text(abns_value)
        alt_names = split_alt_names(abns_value)
        normalized_alt_names = [normalize_text(name) for name in alt_names if normalize_text(name)]
        if not product_name and not normalized_abns and not normalized_alt_names:
            continue

        if product_name == normalized_term:
            exact_matches.append((1.2, record))
            continue
        if normalized_term in product_name:
            exact_matches.append((1.15, record))
            continue
        if normalized_term in normalized_alt_names:
            exact_matches.append((1.1, record))
            continue
        if normalized_abns and normalized_term in normalized_abns:
            exact_matches.append((1.05, record))
            continue

    if exact_matches:
        exact_matches.sort(key=lambda item: (-item[0], get_field(item[1], "product_name"), get_field(item[1], "reg_num")))
        return [record for _, record in exact_matches[:per_term_limit]]
    return []


def normalize_reg_num(value: str) -> str:
    return re.sub(r"[^0-9-]", "", value or "").strip("-")


def search_dump_by_reg_num(reg_num: str, dump_records: list[dict[str, Any]], per_term_limit: int) -> list[dict[str, Any]]:
    normalized_reg = normalize_reg_num(reg_num)
    if not normalized_reg:
        return []

    matches = [
        record
        for record in dump_records
        if normalize_reg_num(get_field(record, "reg_num")) == normalized_reg
    ]
    matches.sort(key=lambda record: (get_field(record, "product_name"), get_field(record, "reg_num")))
    return matches[:per_term_limit]


def fetch_json(params: dict[str, str]) -> dict[str, Any]:
    url = f"{API_URL}?{urlencode(params)}"
    request = Request(url, headers={"User-Agent": "Mozilla/5.0"})
    try:
        with urlopen(request, timeout=60) as response:
            payload = response.read().decode("utf-8")
    except HTTPError as exc:
        raise RuntimeError(f"HTTP error from APPRIL API: {exc.code} {exc.reason}") from exc
    except URLError as exc:
        raise RuntimeError(f"Network error calling APPRIL API: {exc.reason}") from exc

    try:
        data = json.loads(payload)
    except json.JSONDecodeError as exc:
        raise RuntimeError("APPRIL API did not return valid JSON.") from exc

    if not isinstance(data, dict):
        raise RuntimeError("Unexpected APPRIL API response shape.")
    return data


def stringify_value(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def collect_matches(
    input_rows: Iterable[dict[str, str]],
    include_inactive: bool,
    min_score: float,
    per_term_limit: int,
    sleep_seconds: float,
    source: str,
    dump_records: list[dict[str, Any]] | None,
) -> tuple[list[dict[str, Any]], list[CandidateMatch]]:
    summary_rows: list[dict[str, Any]] = []
    all_candidates: list[CandidateMatch] = []

    for input_row in input_rows:
        herbicide_name = input_row["primary_name"]
        search_name = input_row["search_term"]
        epa_number = input_row.get("epa_number", "")
        search_terms = [epa_number] if epa_number else build_search_terms(search_name)
        deduped_records: dict[str, CandidateMatch] = {}

        for term in search_terms:
            if source == "dump":
                records = search_dump_by_reg_num(term, dump_records or [], per_term_limit) if epa_number else search_dump(term, dump_records or [], per_term_limit)
            else:
                records = search_appril(term, include_inactive, per_term_limit, sleep_seconds)
            for record in records:
                score = 2.0 if epa_number else score_record(search_name, record)
                if score < min_score:
                    continue

                key = record_key(record)
                candidate = CandidateMatch(
                    input_name=epa_number or search_name,
                    primary_name=herbicide_name,
                    search_term=term,
                    score=score,
                    record=record,
                )
                prior = deduped_records.get(key)
                if prior is None or candidate.score > prior.score:
                    deduped_records[key] = candidate

        ranked = sorted(
            deduped_records.values(),
            key=lambda item: (-item.score, normalize_text(get_field(item.record, "product_name")), get_field(item.record, "reg_num")),
        )
        best = ranked[0] if ranked else None
        summary_rows.append(build_summary_row(herbicide_name, epa_number or search_name, search_terms, ranked, best))
        all_candidates.extend(ranked)

    return summary_rows, all_candidates


def score_record(input_name: str, record: dict[str, Any]) -> float:
    normalized_input = normalize_text(input_name)
    product_name = normalize_text(get_field(record, "product_name"))
    alt_names = normalize_text(get_field(record, "abns"))

    product_score = similarity(normalized_input, product_name)
    alt_score = max(
        [similarity(normalized_input, normalize_text(name)) for name in split_alt_names(get_field(record, "abns"))]
        or [0.0]
    )

    exact_bonus = 0.15 if product_name == normalized_input else 0.0
    startswith_bonus = 0.08 if product_name.startswith(normalized_input) or normalized_input.startswith(product_name) else 0.0
    contains_bonus = 0.05 if normalized_input and normalized_input in product_name else 0.0

    return round(max(product_score, alt_score * 0.97, similarity(normalized_input, alt_names) * 0.9) + exact_bonus + startswith_bonus + contains_bonus, 4)


def similarity(left: str, right: str) -> float:
    if not left or not right:
        return 0.0
    return SequenceMatcher(None, left, right).ratio()


def normalize_text(text: str) -> str:
    text = text.lower()
    text = re.sub(r"[%/]", " ", text)
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return normalize_spacing(text)


def split_alt_names(value: str) -> list[str]:
    if not value:
        return []
    pieces = re.split(r"\s*\|\s*|\s*;\s*|\s{2,}", value)
    return [piece.strip() for piece in pieces if piece.strip()]


def record_key(record: dict[str, Any]) -> str:
    reg_num = get_field(record, "reg_num")
    product_name = get_field(record, "product_name")
    return f"{reg_num}|{product_name}".strip("|")


def get_field(record: dict[str, Any], field_name: str) -> str:
    value = record.get(field_name, "")
    if value is None:
        return ""
    return str(value).strip()


def build_summary_row(
    herbicide_name: str,
    search_name: str,
    search_terms: list[str],
    ranked: list[CandidateMatch],
    best: CandidateMatch | None,
) -> dict[str, Any]:
    if best is None:
        return {
            "input_herbicide": herbicide_name,
            "input_search_term": search_name,
            "search_terms": " | ".join(search_terms),
            "best_score": "",
            "best_reg_num": "",
            "best_product_name": "",
            "best_alt_brand_names": "",
            "best_active_ingredients": "",
            "best_active_signature": "",
        }

    return {
        "input_herbicide": herbicide_name,
        "input_search_term": search_name,
        "search_terms": " | ".join(search_terms),
        "best_score": best.score,
        "best_reg_num": get_field(best.record, "reg_num"),
        "best_product_name": get_field(best.record, "product_name"),
        "best_alt_brand_names": get_field(best.record, "abns"),
        "best_active_ingredients": get_field(best.record, "ais"),
        "best_active_signature": ingredient_signature(get_field(best.record, "ais")),
    }


def ingredient_signature(active_ingredients: str) -> str:
    text = active_ingredients.lower()
    text = re.sub(r"\b\d+(?:\.\d+)?\s*%", " ", text)
    text = re.sub(r"\b\d{2,7}-\d{2}-\d\b", " ", text)
    text = re.sub(r"\s+", " ", text)
    text = text.strip(" ,;|")
    return text


def parse_individual_active_ingredients(active_ingredients: str) -> list[str]:
    if not active_ingredients:
        return []

    matches = re.finditer(
        r"(?:^|,\s*)([^()]+?)\s*\([^)]*\)\s*-\s*\(([^)]*%?)\)",
        active_ingredients,
    )
    parsed: list[str] = []
    for match in matches:
        ingredient_name = normalize_spacing(match.group(1).strip(" ,;|"))
        percentage = normalize_spacing(match.group(2).strip())
        if ingredient_name and percentage:
            parsed.append(f"{ingredient_name} - {percentage}")

    if parsed:
        return unique_keep_order(parsed)

    cleaned = re.sub(r"\([^)]*\)", "", active_ingredients)
    cleaned = normalize_spacing(cleaned.replace(" ,", ",").strip(" ,;|"))
    return [cleaned] if cleaned else []


def parse_individual_active_ingredient_names(active_ingredients: str) -> list[str]:
    detailed = parse_individual_active_ingredients(active_ingredients)
    names: list[str] = []
    for item in detailed:
        name = re.sub(r"\s*-\s*[^-]+$", "", item).strip(" ,;|")
        name = normalize_spacing(name)
        if name:
            names.append(name)
    return unique_keep_order(names)


def unique_keep_order(values: Iterable[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        cleaned = value.strip()
        if not cleaned or cleaned in seen:
            continue
        seen.add(cleaned)
        result.append(cleaned)
    return result


def write_output(output_file: Path, summary_rows: list[dict[str, Any]], candidates: list[CandidateMatch]) -> None:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Best Matches"

    summary_headers = [
        "Input Herbicide",
        "Input Search Term",
        "Search Terms",
        "Best Score",
        "Best EPA Reg No",
        "Best Product Name",
        "Best Alternate Brand Names",
        "Best Active Ingredients",
        "Best Active Signature",
    ]
    summary_sheet.append(summary_headers)
    for row in summary_rows:
        summary_sheet.append(
            [
                row["input_herbicide"],
                row["input_search_term"],
                row["search_terms"],
                row["best_score"],
                row["best_reg_num"],
                row["best_product_name"],
                row["best_alt_brand_names"],
                row["best_active_ingredients"],
                row["best_active_signature"],
            ]
        )

    exact_groups_sheet = workbook.create_sheet("Active Ingredient Groups")
    exact_groups_sheet.append(
        [
            "Active Ingredients",
            "Herbicide Count",
            "Input Herbicides",
            "EPA Reg Nos",
            "Matched Product Names",
        ]
    )
    for group in build_active_ingredient_groups(summary_rows):
        exact_groups_sheet.append(group)

    individual_groups_sheet = workbook.create_sheet("Individual AI Groups")
    individual_groups_sheet.append(
        [
            "Active Ingredient",
            "Herbicide Count",
            "Input Herbicides",
            "EPA Reg Nos",
            "Matched Product Names",
        ]
    )
    for group in build_individual_ingredient_groups(summary_rows):
        individual_groups_sheet.append(group)

    individual_name_groups_sheet = workbook.create_sheet("Individual AI Names")
    individual_name_groups_sheet.append(
        [
            "Active Ingredient",
            "Herbicide Count",
            "Input Herbicides",
            "EPA Reg Nos",
            "Matched Product Names",
        ]
    )
    for group in build_individual_ingredient_name_groups(summary_rows):
        individual_name_groups_sheet.append(group)

    autosize_columns(summary_sheet)
    autosize_columns(exact_groups_sheet)
    autosize_columns(individual_groups_sheet)
    autosize_columns(individual_name_groups_sheet)
    workbook.save(output_file)


def build_active_ingredient_groups(summary_rows: list[dict[str, Any]]) -> list[list[Any]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in summary_rows:
        active_ingredients = row["best_active_ingredients"]
        if active_ingredients:
            grouped[active_ingredients].append(row)

    output_rows: list[list[Any]] = []
    for active_ingredients, rows in sorted(grouped.items(), key=lambda item: (-len(item[1]), item[0])):
        output_rows.append(
            [
                active_ingredients,
                len(rows),
                "; ".join(unique_keep_order(row["input_herbicide"] for row in rows)),
                "; ".join(unique_keep_order(row["best_reg_num"] for row in rows)),
                "; ".join(unique_keep_order(row["best_product_name"] for row in rows)),
            ]
        )
    return output_rows


def build_individual_ingredient_groups(summary_rows: list[dict[str, Any]]) -> list[list[Any]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in summary_rows:
        for ingredient in parse_individual_active_ingredients(row["best_active_ingredients"]):
            grouped[ingredient].append(row)

    output_rows: list[list[Any]] = []
    for ingredient, rows in sorted(grouped.items(), key=lambda item: (-len(item[1]), item[0])):
        output_rows.append(
            [
                ingredient,
                len(rows),
                "; ".join(unique_keep_order(row["input_herbicide"] for row in rows)),
                "; ".join(unique_keep_order(row["best_reg_num"] for row in rows)),
                "; ".join(unique_keep_order(row["best_product_name"] for row in rows)),
            ]
        )
    return output_rows


def build_individual_ingredient_name_groups(summary_rows: list[dict[str, Any]]) -> list[list[Any]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in summary_rows:
        for ingredient in parse_individual_active_ingredient_names(row["best_active_ingredients"]):
            grouped[ingredient].append(row)

    output_rows: list[list[Any]] = []
    for ingredient, rows in sorted(grouped.items(), key=lambda item: (-len(item[1]), item[0])):
        output_rows.append(
            [
                ingredient,
                len(rows),
                "; ".join(unique_keep_order(row["input_herbicide"] for row in rows)),
                "; ".join(unique_keep_order(row["best_reg_num"] for row in rows)),
                "; ".join(unique_keep_order(row["best_product_name"] for row in rows)),
            ]
        )
    return output_rows


def autosize_columns(worksheet: Any) -> None:
    for column_cells in worksheet.columns:
        values = [len(str(cell.value)) for cell in column_cells if cell.value is not None]
        if not values:
            continue
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(values) + 2, 80)


def default_output_path(input_file: Path) -> Path:
    return input_file.with_name(f"{input_file.stem}_appril_matches.xlsx")


def resolve_lookup_source(source: str, dump_path: Path | None) -> str:
    if source == "dump":
        return "dump"
    if source == "api":
        return "api"
    if dump_path and dump_path.exists():
        return "dump"
    return "api"


def main() -> int:
    args = parse_args()
    input_file = args.input_file.expanduser().resolve()
    if not input_file.exists():
        print(f"Input file not found: {input_file}", file=sys.stderr)
        return 1

    output_file = (args.output_file or default_output_path(input_file)).expanduser().resolve()
    dump_path = args.appril_dump.expanduser().resolve() if args.appril_dump else None

    try:
        input_rows = load_input_rows(input_file, args.sheet, args.column, "Search Term")
        if not input_rows:
            print("No herbicide names were found in the workbook.", file=sys.stderr)
            return 1

        lookup_source = resolve_lookup_source(args.source, dump_path)
        dump_records = load_dump_records(dump_path, args.include_inactive) if lookup_source == "dump" and dump_path else None
        summary_rows, all_candidates = collect_matches(
            input_rows=input_rows,
            include_inactive=args.include_inactive,
            min_score=args.min_score,
            per_term_limit=args.per_term_limit,
            sleep_seconds=args.sleep_seconds,
            source=lookup_source,
            dump_records=dump_records,
        )
        write_output(output_file, summary_rows, all_candidates)
    except Exception as exc:
        print(str(exc), file=sys.stderr)
        return 1

    print(f"Wrote {len(summary_rows)} herbicides to {output_file}")
    print(f"Captured {len(all_candidates)} possible matches in total")
    print(f"Lookup source: {lookup_source}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
